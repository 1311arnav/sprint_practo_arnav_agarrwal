import openpyxl
import os


class ExcelReader:
    """Read test data from Excel files"""

    def __init__(self, file_path):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        self.workbook = openpyxl.load_workbook(file_path)

    def get_sheet(self, sheet_name):
        """Get sheet by name"""
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        return self.workbook[sheet_name]

    @staticmethod
    def read_excel_by_id(file_path, test_case_id, sheet_name="TestData"):
        """
        Read test data from Excel by TestCaseID

        Args:
            file_path (str): Path to Excel file
            test_case_id (str): Test case ID to search for (e.g., 'TC_01')
            sheet_name (str): Sheet name to read from (default: 'TestData')

        Returns:
            dict: Dictionary with column headers as keys and row values as values
                  Returns None if test case ID not found
        """
        try:
            reader = ExcelReader(file_path)
            sheet = reader.get_sheet(sheet_name)

            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]

            # Find row matching TestCaseID
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and row[0] == test_case_id:
                    return {header: value for header, value in zip(headers, row)}

            return None
        except Exception as e:
            print(f"Error reading Excel: {str(e)}")
            return None

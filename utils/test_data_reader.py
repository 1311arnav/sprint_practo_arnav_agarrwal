"""
Test data reader utility for reading Excel files for parametrized tests
"""
import os
import openpyxl
from utils.logger import get_logger

logger = get_logger("test_data_reader")


def get_all_test_rows():
    """Get all test data rows from Excel file (by row number)"""
    testdata_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")
    excel_file = os.path.join(testdata_dir, "TestData_Master.xlsx")
    
    test_rows = []
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook["TestData"]
        
        # Get all rows starting from row 2
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if row[0].value is not None:
                test_rows.append(row_num)
        
        workbook.close()
    except Exception as e:
        logger.error(f"Error reading test rows from Excel: {e}")
    
    return test_rows


def get_test_data_by_row(row_number):
    """Get test data from a specific row in Excel file"""
    testdata_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")
    excel_file = os.path.join(testdata_dir, "TestData_Master.xlsx")
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook["TestData"]
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Get row data
        row_data = sheet[row_number]
        row_values = [cell.value for cell in row_data]
        
        workbook.close()
        
        return {header: value for header, value in zip(headers, row_values)}
    except Exception as e:
        logger.error(f"Error reading test data from row {row_number}: {e}")
        return None


def get_all_tc_02_rows():
    """Get all TC_02 test data rows from Excel file (by row number)"""
    testdata_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")
    excel_file = os.path.join(testdata_dir, "test_data_tc_02.xlsx")
    
    test_rows = []
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # Get all rows starting from row 2
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if row[0].value is not None:
                test_rows.append(row_num)
        
        workbook.close()
    except Exception as e:
        logger.error(f"Error reading TC_02 test rows from Excel: {e}")
    
    return test_rows


def get_test_data_tc_02_by_row(row_number):
    """Get test data for TC_02 from test_data_tc_02.xlsx file"""
    testdata_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")
    excel_file = os.path.join(testdata_dir, "test_data_tc_02.xlsx")
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Get row data
        row_data = sheet[row_number]
        row_values = [cell.value for cell in row_data]
        
        workbook.close()
        
        return {header: value for header, value in zip(headers, row_values) if header is not None}
    except Exception as e:
        logger.error(f"Error reading TC_02 test data from row {row_number}: {e}")
        return None


def get_all_tc_03_rows():
    """Get all TC_03 test data rows from Excel file (by row number)"""
    testdata_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")
    excel_file = os.path.join(testdata_dir, "test_data_tc_03.xlsx")
    
    test_rows = []
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # Get all rows starting from row 2
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if row[0].value is not None:
                test_rows.append(row_num)
        
        workbook.close()
    except Exception as e:
        logger.error(f"Error reading TC_03 test rows from Excel: {e}")
    
    return test_rows


def get_test_data_tc_03_by_row(row_number):
    """Get test data for TC_03 from test_data_tc_03.xlsx file"""
    testdata_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")
    excel_file = os.path.join(testdata_dir, "test_data_tc_03.xlsx")
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Get row data
        row_data = sheet[row_number]
        row_values = [cell.value for cell in row_data]
        
        workbook.close()
        
        return {header: value for header, value in zip(headers, row_values) if header is not None}
    except Exception as e:
        logger.error(f"Error reading TC_03 test data from row {row_number}: {e}")
        return None


def get_all_tc_04_rows():
    """Get all TC_04 test data rows by row number"""
    testdata_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")
    excel_file = os.path.join(testdata_dir, "test_data_tc_04.xlsx")
    
    test_rows = []
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # Get all rows starting from row 2
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if row[0].value is not None:
                test_rows.append(row_num)
        
        workbook.close()
    except Exception as e:
        logger.error(f"Error reading TC_04 test rows from Excel: {e}")
    
    return test_rows


def get_test_data_tc_04_by_row(row_number):
    """Get test data for TC_04 from test_data_tc_04.xlsx file"""
    testdata_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")
    excel_file = os.path.join(testdata_dir, "test_data_tc_04.xlsx")
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Get row data
        row_data = sheet[row_number]
        row_values = [cell.value for cell in row_data]
        
        workbook.close()
        
        return {header: value for header, value in zip(headers, row_values) if header is not None}
    except Exception as e:
        logger.error(f"Error reading TC_04 test data from row {row_number}: {e}")
        return None


def get_all_tc_05_rows():
    """Get all TC_05 test data rows by row number"""
    testdata_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")
    excel_file = os.path.join(testdata_dir, "test_data_tc_05.xlsx")
    
    test_rows = []
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # Get all rows starting from row 2
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if row[0].value is not None:
                test_rows.append(row_num)
        
        workbook.close()
    except Exception as e:
        logger.error(f"Error reading TC_05 test rows from Excel: {e}")
    
    return test_rows


def get_test_data_tc_05_by_row(row_number):
    """Get test data for TC_05 from test_data_tc_05.xlsx file"""
    testdata_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")
    excel_file = os.path.join(testdata_dir, "test_data_tc_05.xlsx")
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Get row data
        row_data = sheet[row_number]
        row_values = [cell.value for cell in row_data]
        
        workbook.close()
        
        return {header: value for header, value in zip(headers, row_values) if header is not None}
    except Exception as e:
        logger.error(f"Error reading TC_05 test data from row {row_number}: {e}")
        return None


def get_all_tc_06_rows():
    """Get all TC_06 test data rows by row number"""
    testdata_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")
    excel_file = os.path.join(testdata_dir, "test_data_tc_06.xlsx")
    
    test_rows = []
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # Get all rows starting from row 2
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if row[0].value is not None:
                test_rows.append(row_num)
        
        workbook.close()
    except Exception as e:
        logger.error(f"Error reading TC_06 test rows from Excel: {e}")
    
    return test_rows


def get_test_data_tc_06_by_row(row_number):
    """Get test data for TC_06 from test_data_tc_06.xlsx file"""
    testdata_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")
    excel_file = os.path.join(testdata_dir, "test_data_tc_06.xlsx")
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Get row data
        row_data = sheet[row_number]
        row_values = [cell.value for cell in row_data]
        
        workbook.close()
        
        return {header: value for header, value in zip(headers, row_values) if header is not None}
    except Exception as e:
        logger.error(f"Error reading TC_06 test data from row {row_number}: {e}")
        return None
